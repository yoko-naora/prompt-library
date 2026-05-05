import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open(r'C:\Users\jding\Desktop\prompt-library\data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

cats = sorted(set(e.get('cat', '') for e in data))
cat_formula = ','.join(cats)

video_missing = []
for i, e in enumerate(data):
    if e.get('type') == 'video' and not e.get('image_prompt', '').strip():
        video_missing.append((i, e.get('link',''), e.get('author',''), e.get('cat','')))

print(f'Categories: {len(cats)}')
print(f'Missing image_prompt: {len(video_missing)}')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "入列"

hfont = Font(name="Arial", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
dfont = Font(name="Arial", size=10)
dalign = Alignment(vertical="center")

headers = ["状态", "类型", "分类", "最终展示URL", "Image Prompt URL", "Video Prompt URL", "源语言", "备注"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = bdr

examples = [
    ["⬜", "中文多推文", "视频制作/Seedance", "https://x.com/MrLarus/status/2051335394550022441", "https://x.com/MrLarus/status/2051002066735071574", "https://x.com/MrLarus/status/2051002488745046029", "CN", "霓裳羽衣舞 - 示例"],
    ["⬜", "英文单推文", "品牌/VI/包装", "https://x.com/AIwithSynthia/status/2051311862265102635", "", "", "EN", "Food of Japan - 示例"],
]
for r, row in enumerate(examples, 2):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = dfont; cell.border = bdr; cell.alignment = dalign

start_row = 4
for j, (idx, link, author, cat) in enumerate(video_missing):
    r = start_row + j
    vals = ["⬜", "中文多推文", cat, "", "", link, "CN", f"#{idx} 补image_prompt"]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = dfont; cell.border = bdr; cell.alignment = dalign

for c, w in enumerate([8, 16, 20, 55, 55, 55, 8, 30], 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

def add_dv(formula, col):
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{start_row + len(video_missing) + 100}")

add_dv(f'"{cat_formula}"', "C")
add_dv('"中文多推文,英文单推文"', "B")
add_dv('"CN,EN"', "G")

ws.freeze_panes = "A2"

# Legend
ws2 = wb.create_sheet("填写说明")
legends = [
    ["列", "说明", "规则"],
    ["状态", "⬜待处理 / ✅已入库", "你不用改，我处理完更新"],
    ["类型", "中文多推文 / 英文单推文", "下拉选择"],
    ["分类", "现有9个分类", "下拉选择"],
    ["最终展示URL", "成果展示推文（图+视频都在的那条）", "必填"],
    ["Image Prompt URL", "GPT-Image2提示词推文(Step1)", "中文多推文必填，英文留空"],
    ["Video Prompt URL", "Seedance提示词推文(Step2)", "中文多推文必填，英文留空"],
    ["源语言", "CN或EN", "翻译方向->JA"],
    ["备注", "额外说明", "可选"],
]
for r, row in enumerate(legends, 1):
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = hfont if r == 1 else dfont
        cell.fill = hfill if r == 1 else PatternFill()
        cell.border = bdr
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 45

wb.save(r"C:\Users\jding\Desktop\prompt_queue.xlsx")
print("Done!")
