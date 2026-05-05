"""
Process prompt_queue.xlsx: delete, update, and add entries to data.json.
"""
import json, re, shutil, subprocess, sys, os
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_FILE = ROOT / "data.json"
IMAGES_DIR = ROOT / "images"
VIDEOS_DIR = ROOT / "videos"
QUEUE_FILE = Path(os.environ["USERPROFILE"]) / "Desktop" / "prompt_queue.xlsx"
GALLERY_DL = "gallery-dl"

def fetch_tweet(url):
    """Fetch tweet content + download media via gallery-dl."""
    tmp = ROOT / ".dl_proc"
    if tmp.exists():
        shutil.rmtree(tmp)
    tmp.mkdir(parents=True)

    subprocess.run([GALLERY_DL, "--no-download", "--write-info-json", "-d", str(tmp), url],
                   capture_output=True, text=True, timeout=90)
    subprocess.run([GALLERY_DL, "-d", str(tmp), url],
                   capture_output=True, text=True, timeout=180)

    # Find JSON
    json_files = list(tmp.rglob("info.json"))
    meta = None
    content = ""
    author = "unknown"
    date = ""
    if json_files:
        with open(json_files[0], "r", encoding="utf-8") as f:
            meta = json.load(f)
        content = meta.get("content", "")
        author = meta.get("author", {}).get("name", "unknown")
        date = meta.get("date", "")

    # Find media
    media_exts = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.mp4', '.webm', '.mov'}
    media_files = [f for f in tmp.rglob("*") if f.is_file() and f.suffix.lower() in media_exts]

    # Move media
    IMAGES_DIR.mkdir(exist_ok=True)
    VIDEOS_DIR.mkdir(exist_ok=True)
    images = []
    video = None
    for mf in media_files:
        ext = mf.suffix.lower()
        dest_name = f"{author}_{mf.name}"
        if ext in ('.mp4', '.webm', '.mov'):
            dest = VIDEOS_DIR / dest_name
            shutil.move(str(mf), str(dest))
            if not video:
                video = f"videos/{dest_name}"
        else:
            dest = IMAGES_DIR / dest_name
            shutil.move(str(mf), str(dest))
            images.append(f"images/{dest_name}")

    if tmp.exists():
        shutil.rmtree(tmp)

    return {"content": content, "author": f"@{author}", "date": date,
            "images": images, "video": video, "meta": meta}


def classify_type(content):
    """Classify as image or video type."""
    if re.search(r'seedance|生成.*视频|生成.*短片|生成.*广告片|视频生成|video prompt', content, re.IGNORECASE):
        return "video"
    return "image"


def extract_prompts(content, ptype):
    """Extract image_prompt and video_prompt from content."""
    image_prompt = ""
    video_prompt = ""

    # Try to split at "Step 1" / "Step 2" markers (Chinese multi-tweet pattern)
    # But for single tweets, the entire content might be one prompt
    if re.search(r'(?i)step\s*[12]', content):
        # Multi-step format
        m = re.search(r'(?is)step\s*1[：:].+?(?=step\s*2[：:]|$)', content)
        if m:
            step1 = re.sub(r'(?i)^step\s*1[：:]\s*', '', m.group(0)).strip()
            # Remove the "先生成..." intro text, keep the prompt starting from "Prompt：" or similar
            prompt_m = re.search(r'(?is)(?:Prompt|提示词|プロンプト)\s*[：:]\s*\n?(.+)', step1)
            if prompt_m:
                image_prompt = prompt_m.group(1).strip()
            else:
                image_prompt = step1

        m = re.search(r'(?is)step\s*2[：:].+?(?=step\s*3[：:]|$)', content)
        if m:
            step2 = re.sub(r'(?i)^step\s*2[：:]\s*', '', m.group(0)).strip()
            prompt_m = re.search(r'(?is)(?:Prompt|提示词|プロンプト)\s*[：:]\s*\n?(.+)', step2)
            if prompt_m:
                video_prompt = prompt_m.group(1).strip()
            else:
                video_prompt = step2
    else:
        # Single tweet: check for image/video prompt markers
        if "gpt-image" in content.lower() or "image prompt" in content.lower() or "image2" in content.lower():
            # Image prompt tweet
            m = re.search(r'(?is)(?:prompt|提示词)\s*[：:]\s*\n?(.+)', content)
            image_prompt = m.group(1).strip() if m else content
        elif "seedance" in content.lower() or "video prompt" in content.lower():
            # Video prompt tweet
            m = re.search(r'(?is)(?:prompt|提示词)\s*[：:]\s*\n?(.+)', content)
            video_prompt = m.group(1).strip() if m else content
        else:
            # Generic: treat whole content as the prompt
            image_prompt = content

    return image_prompt, video_prompt


def main():
    # Load data
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Load Excel
    wb = openpyxl.load_workbook(QUEUE_FILE)
    ws = wb["入列"]

    # Parse rows into actions
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status, stype, cat, final_url, img_url, vid_url, lang, note = row
        if not any([stype, cat, final_url, img_url, vid_url]):
            continue
        status = status or ""
        stype = stype or ""
        cat = cat or ""
        final_url = (final_url or "").strip() if final_url else ""
        img_url = (img_url or "").strip() if img_url else ""
        vid_url = (vid_url or "").strip() if vid_url else ""
        note = (note or "").strip() if note else ""
        rows.append((status, stype, cat, final_url, img_url, vid_url, lang or "CN", note))

    # --- Phase 1: Deletions ---
    to_delete = set()
    for r in rows:
        m = re.search(r'#(\d+)', r[7])
        if m and "删除" in r[7]:
            to_delete.add(int(m.group(1)))

    if to_delete:
        print(f"Deleting {len(to_delete)} entries: {sorted(to_delete)}")
        # Delete from highest index to avoid shifting
        for idx in sorted(to_delete, reverse=True):
            if idx < len(data):
                del data[idx]
                print(f"  Deleted #{idx}")

    # --- Phase 2: Updates ---
    # Now indices have shifted! Need to re-map.
    # We'll find entries by their original tweet link (VID URL)
    # Build a map: VID URL -> current index
    vid_to_idx = {}
    for i, e in enumerate(data):
        link = e.get("link", "")
        if link:
            vid_to_idx[link] = i

    updates = []
    for r in rows:
        m = re.search(r'#(\d+)', r[7])
        if m and "删除" not in r[7] and r[4]:  # Has #number, not delete, has IMG URL
            idx = int(m.group(1))
            updates.append((idx, r[2], r[3], r[4], r[5], r[7]))

    # --- Phase 3: Additions ---
    additions = []
    seen_notes = set()
    for r in rows:
        if not re.search(r'#\d+', r[7]) and r[3] and r[7]:  # No #number, has FINAL URL, has note
            note_key = r[7]
            if note_key in seen_notes:
                continue
            seen_notes.add(note_key)
            additions.append(r)

    print(f"\nUpdates: {len(updates)}")
    print(f"Additions: {len(additions)}")

    # Process additions
    for i, (status, stype, cat, final_url, img_url, vid_url, lang, note) in enumerate(additions):
        print(f"\n--- Adding: {note} ---")
        print(f"  Type: {stype}, Cat: {cat}, Lang: {lang}")
        print(f"  FINAL: {final_url}")
        print(f"  IMG: {img_url}")
        print(f"  VID: {vid_url}")

        if stype == "英文单推文":
            # Single tweet: everything in one URL
            result = fetch_tweet(final_url)
            content = result["content"]
            ptype = classify_type(content)
            image_prompt, video_prompt = extract_prompts(content, ptype)

            entry = {
                "cat": cat,
                "title": result["date"],
                "text": content,
                "link": final_url,
                "title_ja": "",
                "text_ja": "",
                "type": ptype,
                "date": result["date"],
                "author": result["author"],
                "image_prompt": image_prompt,
                "image_prompt_ja": "",
                "video_prompt": video_prompt,
                "video_prompt_ja": "",
                "images": result["images"],
                "video": result["video"] or "",
            }

        else:  # 中文多推文
            # Fetch all 3 tweets
            final_result = fetch_tweet(final_url) if final_url else None
            img_result = fetch_tweet(img_url) if img_url else None
            vid_result = fetch_tweet(vid_url) if vid_url else None

            # Extract prompts
            image_prompt = ""
            video_prompt = ""
            if img_result:
                _, image_prompt = extract_prompts(img_result["content"], "image")
            if vid_result:
                _, video_prompt = extract_prompts(vid_result["content"], "video")

            # Use FINAL for title and media
            title = final_result["date"] if final_result else ""
            author = final_result["author"] if final_result else (img_result["author"] if img_result else "@MrLarus")
            text = final_result["content"] if final_result else ""
            images = final_result["images"] if final_result else []
            video = final_result["video"] if final_result else ""
            link = final_url

            # If no FINAL result media, try VID for video
            if not video and vid_result and vid_result["video"]:
                video = vid_result["video"]
            if not images and img_result:
                images = img_result["images"]

            entry = {
                "cat": cat,
                "title": title,
                "text": text,
                "link": link,
                "title_ja": "",
                "text_ja": "",
                "type": "video",
                "date": title,
                "author": author,
                "image_prompt": image_prompt,
                "image_prompt_ja": "",
                "video_prompt": video_prompt,
                "video_prompt_ja": "",
                "images": images,
                "video": video or "",
            }

        data.append(entry)
        print(f"  Added: type={entry['type']}, images={len(entry['images'])}, video={bool(entry['video'])}")
        print(f"  image_prompt: {len(entry['image_prompt'])} chars")
        print(f"  video_prompt: {len(entry['video_prompt'])} chars")

    # Process updates
    for orig_idx, cat, final_url, img_url, vid_url, note in updates:
        print(f"\n--- Updating #{orig_idx}: {note} ---")

        # Find entry by VID URL
        if vid_url not in vid_to_idx:
            print(f"  WARNING: VID URL not found in data.json: {vid_url}")
            continue
        idx = vid_to_idx[vid_url]
        entry = data[idx]

        # Fetch image_prompt tweet
        if img_url and img_url != vid_url:
            img_result = fetch_tweet(img_url)
            _, image_prompt = extract_prompts(img_result["content"], "image")
            if image_prompt:
                entry["image_prompt"] = image_prompt
                print(f"  Updated image_prompt: {len(image_prompt)} chars")

        # If FINAL URL is different from VID URL, download additional media
        if final_url and final_url != vid_url:
            final_result = fetch_tweet(final_url)
            if final_result["images"]:
                entry["images"].extend(final_result["images"])
            if final_result["video"] and not entry.get("video"):
                entry["video"] = final_result["video"]
            print(f"  Added media from FINAL: +{len(final_result['images'])} images")
            # Use FINAL content as text/title if better
            if final_result["content"]:
                entry["text"] = final_result["content"]
                entry["title"] = final_result["date"]

    # Save
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n=== Done! Total entries: {len(data)} ===")

    # Update Excel statuses
    # Mark processed rows as checked
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        note_cell = row[7]  # 备注 column
        note = (note_cell.value or "").strip() if note_cell.value else ""
        # Mark deleted
        m = re.search(r'#(\d+)', note)
        if m and "删除" in note:
            row[0].value = "✅"
        elif m and not "删除" in note and row[4].value:  # Updated
            row[0].value = "✅"
        elif not re.search(r'#\d+', note) and row[3].value and note:  # New addition
            row[0].value = "✅"

    wb.save(QUEUE_FILE)
    print("Excel status updated.")


if __name__ == "__main__":
    main()
